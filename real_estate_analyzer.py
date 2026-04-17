#!/usr/bin/env python3
"""
Real Estate Investment Analyzer
Generates a formatted Excel workbook with one sheet per property.

Usage:
    python real_estate_analyzer.py
    python real_estate_analyzer.py --output my_portfolio.xlsx
    python real_estate_analyzer.py --fred-key YOUR_KEY   # enables live mortgage rates
                                                          # free key: fred.stlouisfed.org
"""

import sys
import math
import argparse
from datetime import datetime
from dataclasses import dataclass
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing: pip install openpyxl")

try:
    import numpy_financial as npf
except ImportError:
    sys.exit("Missing: pip install numpy-financial")

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# DATA MODEL
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Property:
    # Identity
    name: str
    address: str
    property_type: str = "Single Family"
    year_built: int = 2000
    sqft: int = 1500

    # Purchase
    purchase_price: float = 300_000
    closing_costs_pct: float = 2.5       # % of purchase price
    rehab_costs: float = 0

    # Financing
    down_payment_pct: float = 20         # % of purchase price
    interest_rate: float = 7.0          # annual %
    loan_term_years: int = 30

    # Income
    monthly_rent: float = 2_000
    other_monthly_income: float = 0     # parking, laundry, storage, etc.
    vacancy_rate_pct: float = 5         # %

    # Operating Expenses (annual unless noted)
    property_tax_annual: float = 3_600
    insurance_annual: float = 1_200
    management_fee_pct: float = 8       # % of effective gross income
    maintenance_pct: float = 1.0        # % of property value annually
    hoa_monthly: float = 0
    utilities_monthly: float = 0        # landlord-paid utilities
    other_expenses_monthly: float = 0

    # Projection Assumptions
    appreciation_rate_pct: float = 3.0  # annual property value growth
    rent_growth_rate_pct: float = 2.0   # annual rent growth
    selling_costs_pct: float = 6.0      # % of sale price (agents, fees)
    holding_period_years: int = 10

    # Tax & Depreciation
    land_value_pct: float = 20.0          # % of purchase price that is land (non-depreciable)
    income_tax_rate_pct: float = 32.0     # marginal combined federal + state tax rate
    depreciation_years: float = 27.5      # 27.5 residential / 39 commercial

    # Advanced Parameters
    exit_cap_rate_pct: float = 0.0        # cap-rate exit pricing; 0 = use appreciation model
    capex_reserve_monthly: float = 0.0    # dedicated CAPEX sinking fund (roof, HVAC, etc.)
    mirr_reinvest_rate_pct: float = 6.0   # reinvestment rate assumption for MIRR

    # Optional metadata
    purchase_date: str = ""
    notes: str = ""


# ─────────────────────────────────────────────────────────────────────────────
# FINANCIAL CALCULATIONS
# ─────────────────────────────────────────────────────────────────────────────

class PropertyAnalyzer:
    def __init__(self, prop: Property):
        self.p = prop
        self._run()

    def _run(self):
        p = self.p
        monthly_rate = p.interest_rate / 100 / 12
        n = p.loan_term_years * 12

        # ── Financing ───────────────────────────────────────────
        self.down_payment = p.purchase_price * p.down_payment_pct / 100
        self.closing_costs = p.purchase_price * p.closing_costs_pct / 100
        self.loan_amount = p.purchase_price - self.down_payment
        self.total_cash_invested = self.down_payment + self.closing_costs + p.rehab_costs

        self.loan_amount = max(0.0, self.loan_amount)   # clamp: down > price is allowed

        if n == 0:
            self.monthly_payment = 0.0
        elif monthly_rate > 0:
            self.monthly_payment = (
                self.loan_amount * monthly_rate * (1 + monthly_rate) ** n
                / ((1 + monthly_rate) ** n - 1)
            )
        else:
            self.monthly_payment = self.loan_amount / n
        self.annual_debt_service = self.monthly_payment * 12

        # ── Income ──────────────────────────────────────────────
        self.gross_monthly_income = p.monthly_rent + p.other_monthly_income
        self.gross_annual_income = self.gross_monthly_income * 12
        self.vacancy_loss = self.gross_annual_income * p.vacancy_rate_pct / 100
        self.effective_gross_income = self.gross_annual_income - self.vacancy_loss

        # ── Expenses ────────────────────────────────────────────
        mgmt_fee = self.effective_gross_income * p.management_fee_pct / 100
        maintenance = p.purchase_price * p.maintenance_pct / 100
        pmi_annual = (self.loan_amount * 0.008) if p.down_payment_pct < 20 and self.loan_amount > 0 else 0.0
        capex_annual = p.capex_reserve_monthly * 12

        self.expense_breakdown = {
            "Property Tax":       p.property_tax_annual,
            "Insurance":          p.insurance_annual,
            "Property Management": mgmt_fee,
            "Maintenance & Repairs": maintenance,
            "CAPEX Reserve":      capex_annual,
            "PMI":                pmi_annual,
            "HOA":                p.hoa_monthly * 12,
            "Utilities":          p.utilities_monthly * 12,
            "Other":              p.other_expenses_monthly * 12,
        }
        self.total_operating_expenses = sum(self.expense_breakdown.values())

        # ── NOI & Cash Flow ─────────────────────────────────────
        self.noi = self.effective_gross_income - self.total_operating_expenses
        self.annual_cash_flow = self.noi - self.annual_debt_service
        self.monthly_cash_flow = self.annual_cash_flow / 12

        # ── Key Metrics ─────────────────────────────────────────
        self.cap_rate = (self.noi / p.purchase_price * 100) if p.purchase_price else 0
        self.cash_on_cash = (self.annual_cash_flow / self.total_cash_invested * 100) if self.total_cash_invested else 0
        self.grm = (p.purchase_price / self.gross_annual_income) if self.gross_annual_income else 0
        self.dscr = (self.noi / self.annual_debt_service) if self.annual_debt_service else 0
        self.breakeven_occupancy = (
            (self.total_operating_expenses + self.annual_debt_service)
            / self.gross_annual_income * 100
        ) if self.gross_annual_income else 0
        self.expense_ratio = (self.total_operating_expenses / self.effective_gross_income * 100) if self.effective_gross_income else 0
        self.price_to_rent = p.purchase_price / (p.monthly_rent * 12) if p.monthly_rent else 0
        self.debt_yield = (self.noi / self.loan_amount * 100) if self.loan_amount > 0 else 0

        # ── Amortization Schedule ───────────────────────────────
        self.amortization = self._build_amortization()

        # ── Depreciation & Tax (Year 1) ─────────────────────────
        depreciable_basis = p.purchase_price * (1 - p.land_value_pct / 100)
        self.annual_depreciation = depreciable_basis / p.depreciation_years
        yr1_interest = sum(row["interest"] for row in self.amortization[:12]) if self.amortization else 0
        self.taxable_rental_income = (
            self.effective_gross_income
            - self.total_operating_expenses
            - yr1_interest
            - self.annual_depreciation
        )
        tax_liability_yr1 = max(0.0, self.taxable_rental_income) * p.income_tax_rate_pct / 100
        self.tax_savings_from_depreciation = (
            self.annual_depreciation * p.income_tax_rate_pct / 100
            if self.taxable_rental_income < 0 else 0.0
        )
        self.after_tax_cash_flow = self.annual_cash_flow - tax_liability_yr1
        self.after_tax_monthly_cf = self.after_tax_cash_flow / 12

        # ── Projections & IRR ───────────────────────────────────
        self.projections = self._build_projections()
        self.irr = self._calc_irr()
        self.npv_at_8 = self._calc_npv(0.08)
        self.npv_at_10 = self._calc_npv(0.10)
        self.mirr = self._calc_mirr()
        self.unlevered_irr = self._calc_unlevered_irr()

        # ── Equity Multiple ──────────────────────────────────────
        if self.projections and self.total_cash_invested:
            total_dist = sum(r["annual_cash_flow"] for r in self.projections) + self.projections[-1]["net_sale_proceeds"]
            self.equity_multiple = total_dist / self.total_cash_invested
        else:
            self.equity_multiple = None

        # ── Equity snapshots ────────────────────────────────────
        def equity_at(year):
            idx = min(year * 12 - 1, len(self.amortization) - 1)
            bal = self.amortization[idx]["balance"] if self.amortization else 0
            val = p.purchase_price * (1 + p.appreciation_rate_pct / 100) ** year
            return val, val - bal

        _, self.equity_year1 = equity_at(1)
        _, self.equity_year5 = equity_at(5)
        _, self.equity_year10 = equity_at(10)

    # ── Amortization ────────────────────────────────────────────
    def _build_amortization(self):
        p = self.p
        monthly_rate = p.interest_rate / 100 / 12
        balance = self.loan_amount
        cum_principal = 0.0
        cum_interest = 0.0
        schedule = []
        for month in range(1, p.loan_term_years * 12 + 1):
            interest = balance * monthly_rate
            principal = self.monthly_payment - interest
            balance = max(0.0, balance - principal)
            cum_principal += principal
            cum_interest += interest
            schedule.append({
                "month":                month,
                "payment":              self.monthly_payment,
                "principal":            principal,
                "interest":             interest,
                "balance":              balance,
                "cumulative_principal": cum_principal,
                "cumulative_interest":  cum_interest,
            })
        return schedule

    # ── Year-by-Year Projection ─────────────────────────────────
    def _build_projections(self):
        p = self.p
        rows = []
        rent = p.monthly_rent
        prop_value = p.purchase_price
        cum_cf = 0.0
        cum_after_tax_cf = 0.0

        for year in range(1, p.holding_period_years + 1):
            prop_value *= (1 + p.appreciation_rate_pct / 100)
            if year > 1:
                rent *= (1 + p.rent_growth_rate_pct / 100)

            gross = (rent + p.other_monthly_income) * 12
            vacancy = gross * p.vacancy_rate_pct / 100
            egi = gross - vacancy
            mgmt = egi * p.management_fee_pct / 100
            maint = prop_value * p.maintenance_pct / 100
            capex = p.capex_reserve_monthly * 12
            pmi = (self.loan_amount * 0.008) if p.down_payment_pct < 20 and self.loan_amount > 0 else 0.0
            opex = (p.property_tax_annual + p.insurance_annual + mgmt + maint
                    + capex + pmi
                    + p.hoa_monthly * 12 + p.utilities_monthly * 12 + p.other_expenses_monthly * 12)
            noi = egi - opex
            debt_service = self.annual_debt_service if year <= p.loan_term_years else 0
            cf = noi - debt_service
            cum_cf += cf

            idx = min(year * 12 - 1, len(self.amortization) - 1)
            loan_bal = self.amortization[idx]["balance"] if self.amortization else 0
            equity = prop_value - loan_bal

            # Exit value: cap-rate model if specified, else appreciation model
            if p.exit_cap_rate_pct > 0:
                exit_value = noi / (p.exit_cap_rate_pct / 100)
            else:
                exit_value = prop_value
            net_proceeds = exit_value * (1 - p.selling_costs_pct / 100) - loan_bal
            total_return = cum_cf + net_proceeds - self.total_cash_invested
            roi = total_return / self.total_cash_invested * 100 if self.total_cash_invested else 0

            # After-tax cash flow (simplified: deduct tax on positive taxable income)
            yr_interest = sum(r["interest"] for r in self.amortization[max(0,(year-1)*12):year*12]) if self.amortization else 0
            taxable = egi - opex - yr_interest - self.annual_depreciation
            tax_due = max(0.0, taxable) * p.income_tax_rate_pct / 100
            after_tax_cf = cf - tax_due
            cum_after_tax_cf += after_tax_cf

            # Return on Equity (current year)
            roe = (cf / equity * 100) if equity > 0 else 0.0

            rows.append({
                "year":                 year,
                "prop_value":           prop_value,
                "exit_value":           exit_value,
                "loan_balance":         loan_bal,
                "equity":               equity,
                "noi":                  noi,
                "annual_cash_flow":     cf,
                "after_tax_cash_flow":  after_tax_cf,
                "cumulative_cash_flow": cum_cf,
                "cum_after_tax_cf":     cum_after_tax_cf,
                "net_sale_proceeds":    net_proceeds,
                "total_return":         total_return,
                "roi_if_sold":          roi,
                "roe":                  roe,
                "annual_rent":          gross,
            })
        return rows

    def _future_cash_flows(self):
        """Annual cash flows for years 1..N, with terminal sale proceeds added to year N."""
        flows = []
        for i, row in enumerate(self.projections):
            cf = row["annual_cash_flow"]
            if i == len(self.projections) - 1:
                cf += row["net_sale_proceeds"]
            flows.append(cf)
        return flows

    def _calc_irr(self):
        if not self.projections:
            return None
        try:
            result = npf.irr([-self.total_cash_invested] + self._future_cash_flows())
            if result is None or math.isnan(result) or math.isinf(result):
                return None
            return result * 100
        except Exception:
            return None

    def _calc_npv(self, discount_rate):
        if not self.projections:
            return 0
        try:
            return float(npf.npv(discount_rate, [-self.total_cash_invested] + self._future_cash_flows()))
        except Exception:
            return 0

    def _calc_mirr(self):
        if not self.projections:
            return None
        try:
            result = npf.mirr(
                [-self.total_cash_invested] + self._future_cash_flows(),
                self.p.interest_rate / 100,
                self.p.mirr_reinvest_rate_pct / 100,
            )
            if result is None or math.isnan(result) or math.isinf(result):
                return None
            return result * 100
        except Exception:
            return None

    def _calc_unlevered_irr(self):
        """IRR assuming all-cash purchase — measures property performance without leverage."""
        if not self.projections:
            return None
        try:
            p = self.p
            all_cash_in = (p.purchase_price
                           + p.purchase_price * p.closing_costs_pct / 100
                           + p.rehab_costs)
            flows = []
            for i, row in enumerate(self.projections):
                cf = row["noi"]
                if i == len(self.projections) - 1:
                    cf += row["prop_value"] * (1 - p.selling_costs_pct / 100)
                flows.append(cf)
            result = npf.irr([-all_cash_in] + flows)
            if result is None or math.isnan(result) or math.isinf(result):
                return None
            return result * 100
        except Exception:
            return None


# ─────────────────────────────────────────────────────────────────────────────
# LIVE MORTGAGE RATES (FRED API — optional)
# ─────────────────────────────────────────────────────────────────────────────

def fetch_mortgage_rates(fred_api_key: Optional[str]) -> dict:
    if not REQUESTS_AVAILABLE or not fred_api_key:
        return {}
    series_map = {
        "30-Year Fixed":  "MORTGAGE30US",
        "15-Year Fixed":  "MORTGAGE15US",
        "5/1 ARM":        "MORTGAGE5US",
    }
    rates = {}
    base = "https://api.stlouisfed.org/fred/series/observations"
    for label, sid in series_map.items():
        try:
            r = requests.get(base, params={
                "series_id": sid, "api_key": fred_api_key,
                "file_type": "json", "limit": 1, "sort_order": "desc",
            }, timeout=10)
            obs = r.json().get("observations", [])
            if obs and obs[0].get("value", ".") != ".":
                rates[label] = float(obs[0]["value"])
                rates[f"{label}_date"] = obs[0].get("date", "")
        except Exception:
            pass
    return rates


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

C = {
    "navy":         "1F3864",
    "blue":         "2E75B6",
    "light_blue":   "BDD7EE",
    "very_light":   "DEEAF1",
    "dark_green":   "375623",
    "mid_green":    "70AD47",
    "light_green":  "E2EFDA",
    "yellow":       "FFD966",
    "orange":       "ED7D31",
    "red":          "C00000",
    "light_red":    "FFE0E0",
    "gray":         "595959",
    "light_gray":   "F2F2F2",
    "white":        "FFFFFF",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(cell, *, bold=False, sz=10, fc="000000", italic=False,
       bg=None, ha="left", wrap=False, border=False, nf=None):
    cell.font = _font(bold=bold, size=sz, color=fc, italic=italic)
    cell.alignment = _align(h=ha, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _thin()
    if nf:
        cell.number_format = nf


def _modified_property(prop: Property, **overrides) -> Property:
    """Return a copy of prop with the given field overrides applied."""
    import dataclasses
    d = dataclasses.asdict(prop)
    d.update(overrides)
    return Property(**d)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WORKBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────

class ExcelBuilder:

    def __init__(self, properties: list, fred_key: Optional[str] = None):
        self.properties = properties
        self.analyzers = [PropertyAnalyzer(p) for p in properties]
        self.rates = fetch_mortgage_rates(fred_key)
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def build(self) -> openpyxl.Workbook:
        self._build_summary()
        for prop, analyzer in zip(self.properties, self.analyzers):
            self._build_property_sheet(prop, analyzer)
            self._build_sensitivity_sheet(prop, analyzer)
        return self.wb

    # ── SUMMARY SHEET ─────────────────────────────────────────────────────────

    def _build_summary(self):
        ws = self.wb.create_sheet("Portfolio Summary")
        ws.sheet_properties.tabColor = C["navy"]

        # Title
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value = "REAL ESTATE INVESTMENT PORTFOLIO"
        sc(c, bold=True, sz=18, fc=C["white"], bg=C["navy"], ha="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:M2")
        c = ws["A2"]
        c.value = f"Generated {datetime.now().strftime('%B %d, %Y')}  |  {len(self.properties)} Properties"
        sc(c, italic=True, sz=10, fc=C["gray"], bg=C["light_gray"], ha="center")
        ws.row_dimensions[2].height = 18

        # Column headers row 4
        headers = [
            "Property", "Type", "Purchase Price", "Total Invested",
            "Monthly Rent", "Monthly Cash Flow", "Annual NOI",
            "Cap Rate", "CoC Return", "DSCR", "IRR", "Equity Y10", "Notes"
        ]
        ws.row_dimensions[3].height = 8
        ws.row_dimensions[4].height = 24
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            sc(c, bold=True, sz=10, fc=C["white"], bg=C["blue"], ha="center", border=True)

        totals = dict(pp=0, inv=0, rent=0, mcf=0, noi=0, eq10=0)

        for i, (prop, a) in enumerate(zip(self.properties, self.analyzers)):
            row = 5 + i
            bg = C["white"] if i % 2 == 0 else C["very_light"]
            ws.row_dimensions[row].height = 20

            irr_val = (a.irr / 100) if a.irr is not None else None
            mcf_color = C["mid_green"] if a.monthly_cash_flow >= 0 else C["red"]

            data = [
                (prop.name,              None,          "left"),
                (prop.property_type,     None,          "left"),
                (a.p.purchase_price,     '"$"#,##0',    "right"),
                (a.total_cash_invested,  '"$"#,##0',    "right"),
                (a.p.monthly_rent,       '"$"#,##0',    "right"),
                (a.monthly_cash_flow,    '"$"#,##0.00', "right"),
                (a.noi,                  '"$"#,##0',    "right"),
                (a.cap_rate / 100,       "0.00%",       "right"),
                (a.cash_on_cash / 100,   "0.00%",       "right"),
                (a.dscr,                 "0.00",        "right"),
                (irr_val,                "0.00%",       "right"),
                (a.equity_year10,        '"$"#,##0',    "right"),
                (prop.notes,             None,          "left"),
            ]
            for col, (val, nf, ha) in enumerate(data, 1):
                c = ws.cell(row=row, column=col, value=val)
                sc(c, sz=10, bg=bg, ha=ha, border=True, nf=nf)
                if col == 6:  # monthly cash flow — color coded
                    c.font = Font(size=10, bold=True, color=mcf_color, name="Calibri")

            totals["pp"] += a.p.purchase_price
            totals["inv"] += a.total_cash_invested
            totals["rent"] += a.p.monthly_rent
            totals["mcf"] += a.monthly_cash_flow
            totals["noi"] += a.noi
            totals["eq10"] += a.equity_year10

        # Totals row
        tr = 5 + len(self.properties)
        ws.row_dimensions[tr].height = 22
        total_data = [
            ("TOTALS", None, "left"),
            ("", None, "left"),
            (totals["pp"],   '"$"#,##0',    "right"),
            (totals["inv"],  '"$"#,##0',    "right"),
            (totals["rent"], '"$"#,##0',    "right"),
            (totals["mcf"],  '"$"#,##0.00', "right"),
            (totals["noi"],  '"$"#,##0',    "right"),
            ("", None, "left"),
            ("", None, "left"),
            ("", None, "left"),
            ("", None, "left"),
            (totals["eq10"], '"$"#,##0',    "right"),
            ("", None, "left"),
        ]
        for col, (val, nf, ha) in enumerate(total_data, 1):
            c = ws.cell(row=tr, column=col, value=val)
            sc(c, bold=True, sz=10, fc=C["white"], bg=C["navy"], ha=ha, border=True, nf=nf)

        # Column widths
        widths = [28, 15, 15, 15, 14, 18, 15, 11, 12, 10, 10, 15, 32]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Live rates section
        rates_row = tr + 3
        self._write_rates_block(ws, rates_row, 1)

        ws.freeze_panes = "A5"

    # ── RATES BLOCK ───────────────────────────────────────────────────────────

    def _write_rates_block(self, ws, start_row: int, start_col: int):
        r = start_row
        sc_idx = start_col

        ws.merge_cells(start_row=r, start_column=sc_idx, end_row=r, end_column=sc_idx + 5)
        c = ws.cell(row=r, column=sc_idx,
                    value="CURRENT U.S. MORTGAGE RATES  (Freddie Mac / FRED)" if self.rates
                    else "CURRENT MORTGAGE RATES")
        sc(c, bold=True, sz=11, fc=C["white"],
           bg=C["blue"] if self.rates else C["gray"], ha="center")
        r += 1

        if self.rates:
            for key, val in self.rates.items():
                if key.endswith("_date"):
                    continue
                date_str = self.rates.get(f"{key}_date", "")
                cl = ws.cell(row=r, column=sc_idx, value=key)
                sc(cl, bold=True, sz=10, bg=C["light_blue"], border=True)
                cv = ws.cell(row=r, column=sc_idx + 1, value=val / 100)
                sc(cv, bold=True, sz=10, bg=C["white"], ha="right", border=True, nf="0.00%")
                cd = ws.cell(row=r, column=sc_idx + 2,
                             value=f"Week of {date_str}" if date_str else "")
                sc(cd, italic=True, sz=9, fc=C["gray"], bg=C["white"], border=True)
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=sc_idx, end_row=r, end_column=sc_idx + 5)
            c = ws.cell(row=r, column=sc_idx,
                        value="Add --fred-key YOUR_KEY to enable live rates  |  Free key at fred.stlouisfed.org")
            sc(c, italic=True, sz=9, fc=C["orange"])

    # ── PROPERTY SHEET ────────────────────────────────────────────────────────

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        for ch in r'\/?*[]':
            name = name.replace(ch, "-")
        name = name.replace(":", "-").strip()
        return (name or "Property")[:31]

    def _unique_sheet_name(self, raw: str) -> str:
        base = self._safe_sheet_name(raw)
        if base not in self.wb.sheetnames:
            return base
        i = 2
        while True:
            suffix = f" ({i})"
            candidate = base[:31 - len(suffix)] + suffix
            if candidate not in self.wb.sheetnames:
                return candidate
            i += 1

    # ── SENSITIVITY ANALYSIS SHEET ────────────────────────────────────────────

    def _build_sensitivity_sheet(self, prop: Property, a: PropertyAnalyzer):
        sheet_name = self._unique_sheet_name(f"{self._safe_sheet_name(prop.name)[:20]} - Sensitivity")
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = C["orange"]

        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value = f"SENSITIVITY ANALYSIS  —  {prop.name.upper()}"
        sc(c, bold=True, sz=14, fc=C["white"], bg=C["orange"], ha="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:M2")
        c = ws["A2"]
        c.value = "Green = better than base case  |  Red = worse  |  Values show Annual Cash Flow"
        sc(c, italic=True, sz=9, fc=C["gray"], bg=C["light_gray"], ha="center")
        ws.row_dimensions[2].height = 14

        col_widths = {"A": 28, "B": 4}
        for i in range(3, 14):
            col_widths[get_column_letter(i)] = 14
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        def sens_matrix(start_row, title, row_label, row_deltas, col_label, col_deltas,
                        row_attr, col_attr, metric_fn, cell_nf='"$"#,##0'):
            """Render one NxM sensitivity table."""
            r = start_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(col_deltas))
            c = ws.cell(row=r, column=1, value=title)
            sc(c, bold=True, sz=11, fc=C["white"], bg=C["navy"], ha="center")
            ws.row_dimensions[r].height = 22
            r += 1

            # Column header: col_label
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=2 + len(col_deltas))
            c = ws.cell(row=r, column=3, value=col_label)
            sc(c, bold=True, sz=9, fc=C["white"], bg=C["blue"], ha="center")
            ws.row_dimensions[r].height = 16
            r += 1

            # Column delta headers
            c0 = ws.cell(row=r, column=1, value=row_label)
            sc(c0, bold=True, sz=9, fc=C["white"], bg=C["blue"], ha="center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r + len(row_deltas), end_column=1)
            c1 = ws.cell(row=r, column=2, value="Base →")
            sc(c1, bold=True, sz=8, fc=C["gray"], ha="right")
            for ci, cd in enumerate(col_deltas):
                sign = "+" if cd >= 0 else ""
                c = ws.cell(row=r, column=3 + ci, value=f"{sign}{cd:g}%")
                sc(c, bold=True, sz=9, bg=C["light_blue"], ha="center", border=True)
            ws.row_dimensions[r].height = 18
            r += 1

            base_val = metric_fn(prop, a)
            for ri, rd in enumerate(row_deltas):
                sign = "+" if rd >= 0 else ""
                c = ws.cell(row=r, column=2, value=f"{sign}{rd:g}%")
                sc(c, bold=True, sz=9, bg=C["light_blue"], ha="center", border=True)
                ws.row_dimensions[r].height = 17
                for ci, cd in enumerate(col_deltas):
                    kw = {row_attr: getattr(prop, row_attr) + rd,
                          col_attr: getattr(prop, col_attr) + cd}
                    mod_prop = _modified_property(prop, **kw)
                    mod_a = PropertyAnalyzer(mod_prop)
                    val = metric_fn(mod_prop, mod_a)
                    cell = ws.cell(row=r, column=3 + ci, value=val)
                    if isinstance(val, float):
                        sc(cell, sz=9, ha="right", border=True, nf=cell_nf)
                        diff = val - base_val
                        if abs(diff) > 0.0001:
                            cell.fill = _fill(C["light_green"] if diff > 0 else C["light_red"])
                            cell.font = Font(size=9, bold=True,
                                             color=C["dark_green"] if diff > 0 else C["red"],
                                             name="Calibri")
                    else:
                        sc(cell, sz=9, ha="right", border=True)
                r += 1
            return r + 2

        def annual_cf(p, an):
            return an.annual_cash_flow

        def coc(p, an):
            return an.cash_on_cash / 100   # return as decimal for 0.00% formatting

        row = 4
        # Matrix 1: Vacancy vs Rent change → Annual Cash Flow
        row = sens_matrix(
            row,
            "Annual Cash Flow  —  Vacancy Rate (rows) vs Rent Change (cols)",
            "Vacancy Δ", [-4, -2, 0, 2, 4],
            "Monthly Rent Δ", [-10, -5, 0, 5, 10],
            "vacancy_rate_pct", "monthly_rent",
            annual_cf,
        )
        # Matrix 2: Interest rate vs Down payment → Cash-on-Cash
        row = sens_matrix(
            row,
            "Cash-on-Cash Return  —  Interest Rate Δ (rows) vs Down Payment Δ% (cols)",
            "Rate Δ", [-1.5, -0.5, 0, 0.5, 1.5],
            "Down Payment Δ%", [-5, 0, 5, 10, 15],
            "interest_rate", "down_payment_pct",
            coc,
            cell_nf="0.00%",
        )
        # Matrix 3: Appreciation rate vs Holding period → Total Return
        def total_return_final(p, an):
            return an.projections[-1]["total_return"] if an.projections else 0

        row = sens_matrix(
            row,
            "Total Return at Exit  —  Appreciation Rate Δ (rows) vs Rent Growth Δ (cols)",
            "Apprec. Δ", [-2, -1, 0, 1, 2],
            "Rent Growth Δ", [-1, 0, 1, 2, 3],
            "appreciation_rate_pct", "rent_growth_rate_pct",
            total_return_final,
        )

        ws.freeze_panes = "A4"

    def _build_property_sheet(self, prop: Property, a: PropertyAnalyzer):
        ws = self.wb.create_sheet(self._unique_sheet_name(prop.name))
        ws.sheet_properties.tabColor = C["blue"]

        # Column widths
        col_w = {"A": 30, "B": 18, "C": 3, "D": 30, "E": 18,
                 "F": 3, "G": 28, "H": 18, "I": 3,
                 "J": 14, "K": 16, "L": 16, "M": 16, "N": 16}
        for col, w in col_w.items():
            ws.column_dimensions[col].width = w

        # Title
        ws.merge_cells("A1:N1")
        c = ws["A1"]
        c.value = f"PROPERTY ANALYSIS  —  {prop.name.upper()}"
        sc(c, bold=True, sz=16, fc=C["white"], bg=C["navy"], ha="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:N2")
        c = ws["A2"]
        c.value = (f"{prop.address}   |   {prop.property_type}   |   "
                   f"Built {prop.year_built}   |   {prop.sqft:,} sqft"
                   + (f"   |   {prop.notes}" if prop.notes else ""))
        sc(c, italic=True, sz=10, fc=C["gray"], bg=C["light_gray"], ha="center")
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 8

        # ── Three-column input layout (rows 4+) ──────────────────

        def sec_header(r, col_s, col_e, title, color=C["blue"]):
            ws.merge_cells(start_row=r, start_column=col_s, end_row=r, end_column=col_e)
            c = ws.cell(row=r, column=col_s, value=title)
            sc(c, bold=True, sz=11, fc=C["white"], bg=color, ha="center")
            ws.row_dimensions[r].height = 22

        def data_row(r, lc, vc, label, value, nf=None, bold_val=False, val_color=None):
            cl = ws.cell(row=r, column=lc, value=label)
            sc(cl, sz=10, bg=C["light_gray"], border=True)
            cv = ws.cell(row=r, column=vc, value=value)
            sc(cv, sz=10, bold=bold_val, bg=C["white"], ha="right", border=True, nf=nf)
            if val_color:
                cv.font = Font(size=10, bold=bold_val, color=val_color, name="Calibri")
            ws.row_dimensions[r].height = 18
            return cv

        TOP = 4
        rl, rm, rr = TOP, TOP, TOP

        # ── LEFT: Property Details + Financing ──────────────────
        sec_header(rl, 1, 2, "PROPERTY DETAILS")
        rl += 1
        for label, val, nf in [
            ("Purchase Price",   a.p.purchase_price,      '"$"#,##0'),
            ("Closing Costs",    a.closing_costs,         '"$"#,##0'),
            ("Rehab / CapEx",    a.p.rehab_costs,         '"$"#,##0'),
            ("Down Payment",     a.down_payment,          '"$"#,##0'),
            ("Loan Amount",      a.loan_amount,           '"$"#,##0'),
            ("Total Cash In",    a.total_cash_invested,   '"$"#,##0'),
        ]:
            data_row(rl, 1, 2, label, val, nf, bold_val=(label == "Total Cash In"))
            rl += 1

        rl += 1
        sec_header(rl, 1, 2, "FINANCING")
        rl += 1
        for label, val, nf in [
            ("Interest Rate",       a.p.interest_rate / 100,   "0.00%"),
            ("Loan Term",           f"{a.p.loan_term_years} years", None),
            ("Monthly P&I",         a.monthly_payment,         '"$"#,##0.00'),
            ("Annual Debt Service", a.annual_debt_service,     '"$"#,##0'),
            ("Down Payment %",      a.p.down_payment_pct / 100, "0%"),
        ]:
            data_row(rl, 1, 2, label, val, nf)
            rl += 1

        # ── MIDDLE: Income + Expenses ───────────────────────────
        sec_header(rm, 4, 5, "INCOME (ANNUAL)")
        rm += 1
        for label, val, nf, bold in [
            ("Monthly Rent",          a.p.monthly_rent,            '"$"#,##0',    False),
            ("Other Monthly Income",  a.p.other_monthly_income,    '"$"#,##0',    False),
            ("Gross Annual Income",   a.gross_annual_income,       '"$"#,##0',    False),
            ("Vacancy Rate",          a.p.vacancy_rate_pct / 100,  "0.00%",       False),
            ("Vacancy Loss",          -a.vacancy_loss,             '"$"#,##0',    False),
            ("Effective Gross Income", a.effective_gross_income,   '"$"#,##0',    True),
        ]:
            data_row(rm, 4, 5, label, val, nf, bold_val=bold)
            rm += 1

        rm += 1
        sec_header(rm, 4, 5, "OPERATING EXPENSES (ANNUAL)")
        rm += 1
        for exp_name, exp_val in a.expense_breakdown.items():
            data_row(rm, 4, 5, exp_name, exp_val, '"$"#,##0')
            rm += 1
        data_row(rm, 4, 5, "TOTAL OPERATING EXPENSES", a.total_operating_expenses,
                 '"$"#,##0', bold_val=True)
        rm += 1
        data_row(rm, 4, 5, "Expense Ratio", a.expense_ratio / 100, "0.00%")
        rm += 1

        # ── RIGHT: Key Metrics ──────────────────────────────────
        sec_header(rr, 7, 8, "KEY METRICS DASHBOARD", color=C["navy"])
        rr += 1

        def metric(r, label, value, nf, good=None, bad=None, higher=True):
            cl = ws.cell(row=r, column=7, value=label)
            sc(cl, sz=10, bg=C["light_blue"], border=True)
            cv = ws.cell(row=r, column=8, value=value)
            sc(cv, sz=10, bold=True, bg=C["white"], ha="right", border=True, nf=nf)
            fc = None
            if good is not None and value is not None:
                if higher:
                    fc = (C["mid_green"] if value >= good
                          else C["red"] if bad is not None and value < bad else C["orange"])
                else:
                    fc = (C["mid_green"] if value <= good
                          else C["red"] if bad is not None and value > bad else C["orange"])
            if fc:
                cv.font = Font(size=10, bold=True, color=fc, name="Calibri")
            ws.row_dimensions[r].height = 18

        irr_val = a.irr / 100 if a.irr is not None else None
        mirr_val = a.mirr / 100 if a.mirr is not None else None
        ulirr_val = a.unlevered_irr / 100 if a.unlevered_irr is not None else None
        roi5 = a.projections[4]["roi_if_sold"] / 100 if len(a.projections) >= 5 else None
        roi10 = a.projections[9]["roi_if_sold"] / 100 if len(a.projections) >= 10 else None
        roe_yr1 = a.projections[0]["roe"] / 100 if a.projections else None
        roe_yr5 = a.projections[4]["roe"] / 100 if len(a.projections) >= 5 else None

        metrics_list = [
            # ── Cash Flow ──
            ("Net Operating Income (NOI)",    a.noi,                      '"$"#,##0',    0,    None,  True),
            ("Annual Cash Flow (Pre-Tax)",     a.annual_cash_flow,         '"$"#,##0',    0,    None,  True),
            ("After-Tax Cash Flow (Yr 1)",     a.after_tax_cash_flow,      '"$"#,##0',    0,    None,  True),
            ("Monthly Cash Flow",              a.monthly_cash_flow,        '"$"#,##0.00', 0,    None,  True),
            # ── Returns ──
            ("Cap Rate",                       a.cap_rate / 100,           "0.00%",       0.05, 0.03,  True),
            ("Cash-on-Cash Return (CoC)",      a.cash_on_cash / 100,       "0.00%",       0.08, 0.04,  True),
            ("IRR (Levered)",                  irr_val,                    "0.00%",       0.12, 0.06,  True),
            ("MIRR",                           mirr_val,                   "0.00%",       0.10, 0.05,  True),
            ("Unlevered IRR (All-Cash)",        ulirr_val,                  "0.00%",       0.07, 0.04,  True),
            ("NPV @ 8% Discount Rate",         a.npv_at_8,                 '"$"#,##0',    0,    None,  True),
            ("NPV @ 10% Discount Rate",        a.npv_at_10,                '"$"#,##0',    0,    None,  True),
            ("Equity Multiple",                a.equity_multiple,          "0.00x",       1.5,  1.0,   True),
            # ── Valuation ──
            ("Gross Rent Multiplier (GRM)",    a.grm,                      "0.00",        None, None,  False),
            ("Price-to-Rent Ratio",            a.price_to_rent,            "0.00",        None, None,  False),
            # ── Risk ──
            ("Debt Service Coverage (DSCR)",   a.dscr,                     "0.00",        1.25, 1.0,   True),
            ("Debt Yield",                     a.debt_yield / 100,         "0.00%",       0.08, 0.06,  True),
            ("Break-even Occupancy",           a.breakeven_occupancy/100,  "0.00%",       None, None,  False),
            ("Expense Ratio",                  a.expense_ratio / 100,      "0.00%",       None, None,  False),
            # ── Tax & Depreciation ──
            ("Annual Depreciation",            a.annual_depreciation,      '"$"#,##0',    None, None,  True),
            ("Taxable Rental Income (Yr 1)",   a.taxable_rental_income,    '"$"#,##0',    None, None,  True),
            # ── Equity & ROE ──
            ("Return on Equity Year 1",        roe_yr1,                    "0.00%",       0.06, 0.03,  True),
            ("Return on Equity Year 5",        roe_yr5,                    "0.00%",       0.06, 0.03,  True),
            ("Equity Year 1",                  a.equity_year1,             '"$"#,##0',    0,    None,  True),
            ("Equity Year 5",                  a.equity_year5,             '"$"#,##0',    0,    None,  True),
            ("Equity Year 10",                 a.equity_year10,            '"$"#,##0',    0,    None,  True),
            # ── Exit ──
            ("ROI if Sold Year 5",             roi5,                       "0.00%",       0.15, 0.05,  True),
            ("ROI if Sold Year 10",            roi10,                      "0.00%",       0.30, 0.10,  True),
        ]
        for lbl, val, nf, good, bad, higher in metrics_list:
            metric(rr, lbl, val, nf, good, bad, higher)
            rr += 1

        section_end = max(rl, rm, rr) + 2

        # ── LOAN AMORTIZATION ─────────────────────────────────────────────────

        amort_row = section_end
        ws.row_dimensions[amort_row].height = 8
        amort_row += 1

        sec_header(amort_row, 1, 13, "LOAN AMORTIZATION SCHEDULE", color=C["navy"])
        amort_row += 1
        ws.row_dimensions[amort_row].height = 22

        # Monthly table header (cols A–E, first 60 months)
        for col, h in enumerate(["Month", "Payment", "Principal", "Interest", "Bal. Remaining"], 1):
            c = ws.cell(row=amort_row, column=col, value=h)
            sc(c, bold=True, sz=9, fc=C["white"], bg=C["blue"], ha="center", border=True)

        # Annual summary header (cols G–K)
        for col, h in enumerate(["Year", "Cum. Principal", "Cum. Interest", "Balance", "Equity"], 7):
            c = ws.cell(row=amort_row, column=col, value=h)
            sc(c, bold=True, sz=9, fc=C["white"], bg=C["dark_green"], ha="center", border=True)

        # Monthly rows (first 60 months = 5 years)
        for i, am in enumerate(a.amortization[:60]):
            dr = amort_row + 1 + i
            bg = C["white"] if i % 2 == 0 else C["light_gray"]
            ws.row_dimensions[dr].height = 15
            for col, (val, nf) in enumerate([
                (am["month"],    "0"),
                (am["payment"],  '"$"#,##0.00'),
                (am["principal"], '"$"#,##0.00'),
                (am["interest"], '"$"#,##0.00'),
                (am["balance"],  '"$"#,##0.00'),
            ], 1):
                c = ws.cell(row=dr, column=col, value=val)
                sc(c, sz=9, bg=bg, ha="right", border=True, nf=nf)

        # Annual summary rows (all years)
        for yr in range(1, a.p.loan_term_years + 1):
            idx = yr * 12 - 1
            if idx >= len(a.amortization):
                break
            am = a.amortization[idx]
            dr = amort_row + yr
            ws.row_dimensions[dr].height = 15
            bg = C["white"] if yr % 2 == 0 else C["light_green"]
            prop_val_yr = a.p.purchase_price * (1 + a.p.appreciation_rate_pct / 100) ** yr
            equity_yr = prop_val_yr - am["balance"]
            for col, (val, nf) in enumerate([
                (yr,                         "0"),
                (am["cumulative_principal"],  '"$"#,##0'),
                (am["cumulative_interest"],   '"$"#,##0'),
                (am["balance"],               '"$"#,##0'),
                (equity_yr,                   '"$"#,##0'),
            ], 7):
                c = ws.cell(row=dr, column=col, value=val)
                sc(c, sz=9, bg=bg, ha="right", border=True, nf=nf)

        amort_end = amort_row + max(60, a.p.loan_term_years) + 3

        # ── YEAR-BY-YEAR PROJECTION ───────────────────────────────────────────

        proj_row = amort_end
        ws.row_dimensions[proj_row].height = 8
        proj_row += 1

        sec_header(proj_row, 1, 13, f"{a.p.holding_period_years}-YEAR INVESTMENT PROJECTION", color=C["dark_green"])
        proj_row += 1
        ws.row_dimensions[proj_row].height = 22

        proj_heads = ["Year", "Property Value", "Loan Balance", "Equity",
                      "Annual NOI", "Pre-Tax CF", "After-Tax CF", "Cum. CF",
                      "Net Sale Proceeds", "Total Return", "ROI if Sold",
                      "ROE", "Exit Value"]
        for col, h in enumerate(proj_heads, 1):
            c = ws.cell(row=proj_row, column=col, value=h)
            sc(c, bold=True, sz=9, fc=C["white"], bg=C["mid_green"], ha="center", border=True)

        for i, prow in enumerate(a.projections):
            dr = proj_row + 1 + i
            bg = C["white"] if i % 2 == 0 else C["light_green"]
            ws.row_dimensions[dr].height = 17
            vals = [
                (prow["year"],                  "0"),
                (prow["prop_value"],             '"$"#,##0'),
                (prow["loan_balance"],           '"$"#,##0'),
                (prow["equity"],                 '"$"#,##0'),
                (prow["noi"],                    '"$"#,##0'),
                (prow["annual_cash_flow"],       '"$"#,##0'),
                (prow["after_tax_cash_flow"],    '"$"#,##0'),
                (prow["cumulative_cash_flow"],   '"$"#,##0'),
                (prow["net_sale_proceeds"],      '"$"#,##0'),
                (prow["total_return"],           '"$"#,##0'),
                (prow["roi_if_sold"] / 100,      "0.00%"),
                (prow["roe"] / 100,              "0.00%"),
                (prow["exit_value"],             '"$"#,##0'),
            ]
            for col, (val, nf) in enumerate(vals, 1):
                c = ws.cell(row=dr, column=col, value=val)
                sc(c, sz=9, bg=bg, ha="right", border=True, nf=nf)
                if col in (6, 7, 10):
                    col_r = C["mid_green"] if val >= 0 else C["red"]
                    c.font = Font(size=9, bold=True, color=col_r, name="Calibri")

        proj_end = proj_row + len(a.projections) + 3

        # ── LIVE RATES ────────────────────────────────────────────────────────

        self._write_rates_block(ws, proj_end + 1, 1)

        ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# SAMPLE PROPERTIES
# ─────────────────────────────────────────────────────────────────────────────

SAMPLE_PROPERTIES = [
    Property(
        name="123 Main St",
        address="123 Main Street, Austin, TX 78701",
        property_type="Single Family",
        year_built=2005,
        sqft=1850,
        purchase_price=350_000,
        closing_costs_pct=2.5,
        rehab_costs=5_000,
        down_payment_pct=20,
        interest_rate=7.25,
        loan_term_years=30,
        monthly_rent=2_400,
        vacancy_rate_pct=5,
        property_tax_annual=4_200,
        insurance_annual=1_500,
        management_fee_pct=8,
        maintenance_pct=1.0,
        capex_reserve_monthly=200,
        appreciation_rate_pct=4.0,
        rent_growth_rate_pct=3.0,
        holding_period_years=10,
        land_value_pct=20,
        income_tax_rate_pct=32,
        exit_cap_rate_pct=5.5,
        notes="Tech corridor — strong rental demand",
    ),
    Property(
        name="456 Oak Ave",
        address="456 Oak Avenue, Nashville, TN 37201",
        property_type="Duplex",
        year_built=1998,
        sqft=2_400,
        purchase_price=420_000,
        closing_costs_pct=2.5,
        rehab_costs=15_000,
        down_payment_pct=25,
        interest_rate=7.0,
        loan_term_years=30,
        monthly_rent=3_200,
        other_monthly_income=100,
        vacancy_rate_pct=7,
        property_tax_annual=3_800,
        insurance_annual=1_800,
        management_fee_pct=8,
        maintenance_pct=1.2,
        capex_reserve_monthly=300,
        appreciation_rate_pct=3.5,
        rent_growth_rate_pct=2.5,
        holding_period_years=10,
        land_value_pct=15,
        income_tax_rate_pct=35,
        exit_cap_rate_pct=5.0,
        notes="Both units leased, long-term tenants",
    ),
    Property(
        name="789 Pine Rd",
        address="789 Pine Road, Phoenix, AZ 85001",
        property_type="Single Family",
        year_built=2015,
        sqft=2_100,
        purchase_price=290_000,
        closing_costs_pct=2.0,
        rehab_costs=0,
        down_payment_pct=20,
        interest_rate=7.5,
        loan_term_years=30,
        monthly_rent=2_100,
        vacancy_rate_pct=4,
        property_tax_annual=2_600,
        insurance_annual=1_100,
        management_fee_pct=10,
        maintenance_pct=0.8,
        capex_reserve_monthly=150,
        appreciation_rate_pct=5.0,
        rent_growth_rate_pct=3.0,
        holding_period_years=10,
        land_value_pct=25,
        income_tax_rate_pct=28,
        exit_cap_rate_pct=6.0,
        notes="New construction, low near-term maintenance",
    ),
]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Real Estate Investment Analyzer — generates an Excel workbook"
    )
    parser.add_argument("--output",   default="real_estate_analysis.xlsx",
                        help="Output file path (default: real_estate_analysis.xlsx)")
    parser.add_argument("--fred-key", default=None,
                        help="FRED API key for live mortgage rates (free at fred.stlouisfed.org)")
    args = parser.parse_args()

    print(f"\nReal Estate Investment Analyzer")
    print(f"{'─' * 40}")
    print(f"Properties: {len(SAMPLE_PROPERTIES)}")
    if args.fred_key:
        print(f"Fetching live mortgage rates from FRED...")
    else:
        print(f"Live rates: disabled (pass --fred-key to enable)")
    print()

    builder = ExcelBuilder(SAMPLE_PROPERTIES, fred_key=args.fred_key)
    wb = builder.build()
    wb.save(args.output)

    print(f"✓  Saved → {args.output}")
    print()
    print("Metrics per property:")
    metrics = [
        "NOI, Cap Rate, Cash-on-Cash Return",
        "Monthly & Annual Cash Flow",
        "IRR, NPV (8% & 10% discount rates)",
        "Gross Rent Multiplier (GRM)",
        "Debt Service Coverage Ratio (DSCR)",
        "Price-to-Rent Ratio, Expense Ratio",
        "Break-even Occupancy Rate",
        "ROI if sold at Year 5 / Year 10",
        "Equity build-up at Year 1 / 5 / 10",
        "Full loan amortization (30 years)",
        f"Year-by-year projection (appreciation, rent growth, sale proceeds)",
    ]
    for m in metrics:
        print(f"  • {m}")
    print()
    if not args.fred_key:
        print("Tip: get free FRED API key at https://fred.stlouisfed.org/docs/api/api_key.html")
        print("     then run:  python real_estate_analyzer.py --fred-key YOUR_KEY")
    print()


if __name__ == "__main__":
    main()
